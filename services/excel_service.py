# ===========================================================
# Módulo: services/excel_service.py
# Función: Exportar registros de asistencia a un archivo Excel
#          y generar la plantilla de carga de trabajadores.
# Funciones:
#   - exportar_asistencia(registros, ruta)
#   - generar_plantilla_trabajadores(ruta)
# Dependencias: openpyxl, os, config
# ===========================================================

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import Config

# Estilo de encabezado reutilizable
_ENCABEZADO_FUENTE = Font(bold=True, color="FFFFFF")
_ENCABEZADO_FONDO = PatternFill("solid", fgColor="1F3A5F")
_CENTRO = Alignment(horizontal="center", vertical="center")


def _aplicar_encabezado(ws, encabezados):
    """Escribe la fila de encabezados con estilo en la hoja dada."""
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = _ENCABEZADO_FUENTE
        celda.fill = _ENCABEZADO_FONDO
        celda.alignment = _CENTRO


def exportar_asistencia(registros, nombre_archivo="CONTROL_ASISTENCIA.xlsx"):
    """Crea un Excel con los registros de asistencia recibidos.
    Devuelve la ruta absoluta del archivo generado."""
    os.makedirs(Config.EXCEL_FOLDER, exist_ok=True)
    ruta = os.path.join(Config.EXCEL_FOLDER, nombre_archivo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    _aplicar_encabezado(
        ws,
        ["Documento", "Nombres", "Apellidos", "Fecha",
         "Entrada", "Salida", "Observación"],
    )
    for r in registros:
        ws.append([
            r.get("documento"), r.get("nombres"), r.get("apellidos"),
            r.get("fecha"), r.get("hora_entrada"), r.get("hora_salida"),
            r.get("observacion"),
        ])

    # Ancho de columnas automático aproximado
    for col in ws.columns:
        ancho = max((len(str(c.value)) if c.value else 0) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = ancho

    wb.save(ruta)
    return ruta


def exportar_concentracion(filas, rango_inicio, rango_fin,
                           nombre_archivo="CONCENTRACION_POR_SITIO.xlsx"):
    """Crea un Excel con la concentración de personal por sitio
    (presentes, entradas, salidas y total) para el rango dado.
    Devuelve la ruta absoluta del archivo generado."""
    os.makedirs(Config.EXCEL_FOLDER, exist_ok=True)
    ruta = os.path.join(Config.EXCEL_FOLDER, nombre_archivo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Concentración"

    # Título con el rango analizado (en negrilla, sin fondo de encabezado)
    ws.append([f"Concentración por estación  ·  del {rango_inicio} al {rango_fin}"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])  # fila en blanco

    # Encabezado de la tabla en la fila 3. Antes se usaba
    # _aplicar_encabezado, que siempre estiliza la fila 1, y el
    # título quedaba pintado como encabezado por error.
    ws.append(["Estación / Sitio", "Presentes ahora", "Entradas", "Salidas", "Total marcajes"])
    for celda in ws[3]:
        celda.font = _ENCABEZADO_FUENTE
        celda.fill = _ENCABEZADO_FONDO
        celda.alignment = _CENTRO

    total_presentes = 0
    total_marcajes = 0
    for s in filas:
        ws.append([s["sitio"], s["presentes"], s["entradas"],
                   s["salidas"], s["total"]])
        total_presentes += s["presentes"]
        total_marcajes += s["total"]

    # Fila de totales
    fila_total = ["TOTAL", total_presentes, "", "", total_marcajes]
    ws.append(fila_total)
    for celda in ws[ws.max_row]:
        celda.font = Font(bold=True)

    # Ancho de columnas automático aproximado
    for col in ws.columns:
        ancho = max((len(str(c.value)) if c.value else 0) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = ancho

    wb.save(ruta)
    return ruta


def generar_plantilla_trabajadores(nombre_archivo="PLANTILLA_TRABAJADORES.xlsx"):
    """Genera una plantilla vacía para cargar trabajadores masivamente."""
    os.makedirs(Config.EXCEL_FOLDER, exist_ok=True)
    ruta = os.path.join(Config.EXCEL_FOLDER, nombre_archivo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Trabajadores"
    _aplicar_encabezado(
        ws, ["Documento", "Nombres", "Apellidos", "Cargo", "Área", "Turno"]
    )
    # Fila de ejemplo para guiar al usuario (la puede borrar)
    ws.append(["1012345678", "Carlos", "Ramírez", "Operario", "Obra", "Diurno"])
    wb.save(ruta)
    return ruta


def leer_trabajadores_excel(ruta_archivo):
    """Lee un archivo Excel de trabajadores y devuelve una lista de
    diccionarios listos para insertar. Tolera columnas en cualquier
    orden siempre que los encabezados coincidan (sin importar mayúsculas
    ni tildes en 'Área').

    Devuelve (filas_validas, errores) donde:
      - filas_validas: lista de dicts con las claves del modelo
      - errores: lista de textos describiendo filas con problemas
    """
    from openpyxl import load_workbook

    # Mapa de encabezados aceptados -> clave interna
    mapa = {
        "documento": "documento",
        "nombres": "nombres",
        "apellidos": "apellidos",
        "cargo": "cargo",
        "area": "area", "área": "area",
        "turno": "turno",
    }

    # Si el archivo no es un .xlsx válido (corrupto, renombrado, etc.),
    # load_workbook lanza una excepción y antes tumbaba la petición (500).
    try:
        wb = load_workbook(ruta_archivo, read_only=True, data_only=True)
    except Exception:
        return [], ["El archivo no es un Excel .xlsx válido. "
                    "Descargue la plantilla y vuelva a intentarlo."]
    ws = wb.active

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return [], ["El archivo está vacío."]

    # La primera fila son los encabezados; se normalizan a minúsculas
    encabezados = [str(c).strip().lower() if c else "" for c in filas[0]]
    indices = {}
    for i, enc in enumerate(encabezados):
        if enc in mapa:
            indices[mapa[enc]] = i

    # Documento, nombres y apellidos son obligatorios
    for obligatorio in ("documento", "nombres", "apellidos"):
        if obligatorio not in indices:
            return [], [f"Falta la columna obligatoria '{obligatorio}'. "
                        "Use la plantilla descargable."]

    validas, errores = [], []
    for n, fila in enumerate(filas[1:], start=2):  # n = número de fila en Excel
        # Saltar filas totalmente vacías
        if not any(c is not None and str(c).strip() != "" for c in fila):
            continue

        def celda(clave):
            i = indices.get(clave)
            if i is None or i >= len(fila) or fila[i] is None:
                return ""
            return str(fila[i]).strip()

        documento = celda("documento")
        nombres = celda("nombres")
        apellidos = celda("apellidos")

        if not documento or not nombres or not apellidos:
            errores.append(f"Fila {n}: faltan documento, nombres o apellidos.")
            continue

        validas.append({
            "documento": documento,
            "nombres": nombres,
            "apellidos": apellidos,
            "cargo": celda("cargo"),
            "area": celda("area"),
            "turno": celda("turno"),
        })

    return validas, errores
